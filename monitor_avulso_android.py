# -*- coding: utf-8 -*-
"""
Expediente ALECE — Android/Termux (Requests + BS4)
Versão integrada com OneDrive:
- Faz download da planilha antes de ler (rclone copy --update -v)
- Faz upload após registrar novas publicações
- Exibe logs do rclone em tempo real
- Só salva se existir a palavra-chave (ex: "mensagem") e pelo menos um número
"""

import os, re, shlex, subprocess, unicodedata
from time import sleep
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# ======== CONFIGURAÇÕES ========
URL = "https://www.al.ce.gov.br/legislativo/ordem-do-dia/avulso-de-projeto"
HEADERS = {"User-Agent": "Mozilla/5.0 (Linux; Android 13) Chrome Mobile Safari/537.36"}
BASE_DIR = "/data/data/com.termux/files/home/storage/documents/escaner"
PASTA_PDFS = os.path.join(BASE_DIR, "mensagens"); os.makedirs(PASTA_PDFS, exist_ok=True)
ARQ_EXCEL = os.path.join(BASE_DIR, "mensagens_encontradas_avulso.xlsx")
ABA_EXCEL = "encontradas"
NUMEROS_DESTINO = ["558588227227"]
# NUMEROS_DESTINO = ["558588227227", "558597159955", "558587262526", "558596195560, 558581645454"]
SENDER_CANDIDATOS = ["/storage/emulated/0/Documents/escaner/sender_baileys.js"]

# ======== PALAVRA-CHAVE ALTERÁVEL ========
# Basta trocar o valor abaixo por "projeto", "decreto", "indicação", etc.
PALAVRA_CHAVE = "mensagem"

# ======== EXPRESSÕES REGULARES ========
RX_NUM_ANO = re.compile(r"\b(\d{1,5})/(\d{4})\b")  # ex: 85/2025
RX_MENSAGEM = re.compile(r"mensagem\s*n[º°]?\s*\.?\s*(\d{1,6})", re.IGNORECASE)

# ======== SINCRONIZAÇÃO COM ONEDRIVE ========
def executar_rclone(comando: str):
    print("▶️ Executando:", comando)
    proc = subprocess.Popen(comando, shell=True, text=True)
    proc.communicate()
    if proc.returncode == 0:
        print("✅ rclone finalizado com sucesso.\n")
    else:
        print(f"⚠️ rclone terminou com código {proc.returncode}.\n")

def baixar():
    remote_dir = "onedrive:/Gabinete/site/listas"
    print("📥 Baixando planilha do OneDrive...")
    executar_rclone(f"rclone copy {remote_dir} {BASE_DIR} --update -v")
    print("✅ Baixa finalizada.\n")

def upload():
    remote_dir = "onedrive:/Gabinete/site/listas"
    print("📤 Enviando planilha para OneDrive...")
    executar_rclone(f"rclone copy {ARQ_EXCEL} {remote_dir} --update -v")
    print("✅ Upload finalizado.\n")

# ======== FUNÇÕES DE APOIO ========
def normalize(s: str) -> str:
    if not s: return ""
    t = unicodedata.normalize("NFD", s)
    return "".join(c for c in t if unicodedata.category(c) != "Mn").lower()

def excel_init():
    p = Path(ARQ_EXCEL)
    if p.exists():
        wb = load_workbook(ARQ_EXCEL)
        ws = wb[ABA_EXCEL] if ABA_EXCEL in wb.sheetnames else wb.create_sheet(ABA_EXCEL)
    else:
        wb = Workbook(); ws = wb.active
        ws.title = ABA_EXCEL; ws.append(["chave", "data"])
        wb.save(ARQ_EXCEL)
    return wb, ws

def enviados_carregar() -> set:
    wb, ws = excel_init(); out = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0]:
            out.add(str(row[0]).strip())
    return out

def enviados_salvar(chave: str):
    wb, ws = excel_init()
    ws.append([chave, datetime.now().strftime("%d/%m/%Y %H:%M")])
    wb.save(ARQ_EXCEL)

def localizar_sender_js() -> Tuple[str, str]:
    for caminho in SENDER_CANDIDATOS:
        if os.path.isfile(caminho):
            return caminho, os.path.dirname(caminho)
    raise FileNotFoundError("sender_baileys.js não encontrado nas pastas padrão.")

def checar_node():
    try:
        out = subprocess.run(["node", "-v"], capture_output=True, text=True)
        if out.returncode != 0:
            raise RuntimeError(out.stderr.strip() or "Node indisponível")
    except FileNotFoundError:
        raise RuntimeError("Instale Node: pkg install -y nodejs-lts")

def enviar_mensagem(numeros: List[str], mensagem: str, caminho_pdf: Optional[str] = None):
    checar_node()
    sender_js, cwd = localizar_sender_js()
    args = ["node", sender_js, ",".join(numeros), mensagem]
    if caminho_pdf and os.path.isfile(caminho_pdf) and os.path.getsize(caminho_pdf) > 0:
        args.append(caminho_pdf)
    print("▶️ Enviando:", " ".join(shlex.quote(a) for a in args))
    proc = subprocess.Popen(args, cwd=cwd)
    proc.wait()
    if proc.returncode != 0:
        print(f"⚠️ enviar_mensagem.js saiu com código {proc.returncode}")

def download_pdf(url, download_dir):
    from urllib.parse import urlparse
    try:
        r = requests.get(url, headers=HEADERS, timeout=60)
        if r.status_code != 200:
            print(f"❌ HTTP {r.status_code} ao baixar {url}")
            return None
        os.makedirs(download_dir, exist_ok=True)
        path = urlparse(url).path
        parts = path.strip("/").split("/")
        arq = parts[-1]
        dest = os.path.join(download_dir, arq)
        with open(dest, "wb") as f:
            f.write(r.content)
        if os.path.getsize(dest) <= 0:
            print("⚠️ PDF 0B:", dest)
            return None
        print("✅ Baixado:", dest)
        return dest
    except Exception as e:
        print("⚠️ Exceção ao baixar:", e)
        return None

# ======== COLETA DE LINKS ========
def coletar_mensagens() -> List[Dict]:
    r = requests.get(URL, headers=HEADERS, timeout=60)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    itens = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href.lower().endswith(".pdf"):
            continue

        titulo = a.get_text(" ", strip=True)
        texto = titulo + " " + a.parent.get_text(" ", strip=True)

        # Verifica se contém a palavra-chave configurada
        if PALAVRA_CHAVE not in normalize(texto):
            continue

        # Extrai números principais e de "mensagem"
        m_num = RX_NUM_ANO.search(texto)
        m_msg = RX_MENSAGEM.search(texto)
        numero, ano, msg_num = None, None, None

        if m_num:
            numero, ano = m_num.group(1), m_num.group(2)
        if m_msg:
            msg_num = m_msg.group(1)

        # Só continua se houver pelo menos um número (principal ou mensagem)
        if not (numero or msg_num):
            continue

        # Define a chave única
        if numero and ano and msg_num:
            chave = f"{numero}/{ano}_MSG{msg_num}"
        elif numero and ano:
            chave = f"{numero}/{ano}"
        elif msg_num:
            chave = f"MSG{msg_num}"

        descricao = ""
        base = a.parent if getattr(a.parent, "name", None) == "b" else a
        for sib in base.next_siblings:
            if getattr(sib, "name", None) in {"b", "a"}:
                break
            if getattr(sib, "name", None) == "br":
                continue
            txt = BeautifulSoup(str(sib), "html.parser").get_text(" ", strip=True)
            if txt:
                descricao = txt
                break

        itens.append({
            "numero": numero,
            "ano": ano,
            "msg_num": msg_num,
            "titulo": titulo,
            "descricao": descricao,
            "pdf_url": href,
            "chave": chave
        })

    return itens

# ======== LOOP PRINCIPAL ========
def main_loop():
    while True:
        try:
            baixar()
            enviados = enviados_carregar()
            itens = coletar_mensagens()
            print(f"🔎 Encontrados {len(itens)} links contendo '{PALAVRA_CHAVE}'")

            for it in itens:
                chave = it["chave"]
                if chave in enviados:
                    print("🔁 Já enviado:", chave)
                    continue

                mensagem = f"Votação do seguinte Projeto: {it['titulo']} {it['descricao']}".strip()
                caminho_pdf = download_pdf(it["pdf_url"], PASTA_PDFS)
                enviar_mensagem(NUMEROS_DESTINO, mensagem, caminho_pdf)
                enviados_salvar(chave)
                upload()
                print("✅ Enviado e registrado:", chave)
                sleep(2)

        except KeyboardInterrupt:
            print("\nInterrompido.")
            return
        except Exception as e:
            print("Erro no loop:", e)

        hh = datetime.now().strftime("%H:%M:%S")
        print(f"⏰ Horário: {hh} — dormindo 1800s\n")
        sleep(1800)

if __name__ == "__main__":
    main_loop()
