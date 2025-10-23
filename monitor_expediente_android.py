# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
"""
Expediente ALECE — Android/Termux (Requests + BS4)
- Varre https://www.al.ce.gov.br/legislativo/expediente
- Coleta TODOS os <a href> cujo texto contenha "Mensagem"
  (funciona para <b><a>…</a></b> e para <a><b>…</b></a>)
- Extrai numero/ano (####/####) e descrição após o link
- Baixa o PDF diretamente do href
- Envia via sender_baileys.js (Node) com/sem anexo
- Registra no Excel para não duplicar

Dependências (uma vez):
  pkg update -y && pkg install -y python nodejs-lts
  pip install --upgrade pip
  pip install requests beautifulsoup4 openpyxl
  termux-setup-storage

Pasta alvo:
  /storage/emulated/0/Documents/escaner
    ├─ monitor_expediente_android.py   (este arquivo)
    ├─ sender_baileys.js               (seu sender, com node_modules nesta pasta)
    └─ mensagens/                      (PDFs baixados)

Dica: teste o sender sozinho antes:
  cd /storage/emulated/0/Documents/escaner
  node sender_baileys.js "5585SEUNUMERO" "teste"
"""

import os, re, shlex, subprocess, unicodedata
from time import sleep
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

URL = "https://www.al.ce.gov.br/legislativo/expediente"
HEADERS = {"User-Agent":"Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 (KHTML, like Gecko) Chrome Mobile Safari/537.36"}

BASE_DIR   = r"C:\Users\FABIO\OneDrive\Gabinete\site\listas"
PASTA_PDFS = os.path.join(BASE_DIR, "mensagens"); os.makedirs(PASTA_PDFS, exist_ok=True)
ARQ_EXCEL  = os.path.join(BASE_DIR, "mensagens_encontradas.xlsx"); ABA_EXCEL="encontradas"
#NUMEROS_DESTINO2 = ["558588227227"]
NUMEROS_DESTINO2 = ["558588227227", "558597159955", "558596195560"]
#NUMEROS_DESTINO = ["558588227227"]
NUMEROS_DESTINO = ["558588227227", "558597159955", "558587262526", "558596195560"]
# SENDER_CANDIDATOS = [os.path.join(BASE_DIR, "enviar_mensagem.js"),
#                       "/data/data/com.termux/files/home/bot/enviar_mensagem.js"]
SENDER_CANDIDATOS = [
    r"C:\Users\FABIO\OneDrive\Gabinete\site\enviar_mensagem.js",                 # ~/bot
    "/storage/emulated/0/Documents/escaner/sender_baileys.js",               # sdcard/Documentos/escaner
] 
RX_NUM_ANO   = re.compile(r"\b(\d{2,5})/(\d{2,4})\b")
RX_NUMERO2   = re.compile(r"\b(\d{1,2}\.\d{3}|\d{3,5})\b")

def normalize(s:str)->str:
    if not s: return ""
    t = unicodedata.normalize("NFD", s)
    return "".join(c for c in t if unicodedata.category(c)!="Mn").lower()

def tem_mensagem(txt:str)->bool:
    return "mensagem" in normalize(txt)

# ---- Excel
def excel_init():
    p=Path(ARQ_EXCEL)
    if p.exists():
        wb=load_workbook(ARQ_EXCEL)
        ws=wb[ABA_EXCEL] if ABA_EXCEL in wb.sheetnames else wb.create_sheet(ABA_EXCEL)
    else:
        wb=Workbook(); ws=wb.active; ws.title=ABA_EXCEL; ws.append(["numero"]); wb.save(ARQ_EXCEL)
    return wb,ws
def enviados_carregar()->set:
    wb,ws=excel_init(); out=set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0]: out.add(str(row[0]).strip())
    return out
def enviados_salvar(num_ano:str):
    wb,ws=excel_init(); ws.append([num_ano]); wb.save(ARQ_EXCEL)

# ---- sua função de download (baixa diretamente o href do <a>)
def download_pdf(url, download_dir):
    from urllib.parse import urlparse
    try:
        r=requests.get(url, headers=HEADERS, timeout=60)
        if r.status_code!=200:
            print(f"❌ HTTP {r.status_code} ao baixar {url}"); return None
        os.makedirs(download_dir, exist_ok=True)
        path=urlparse(url).path  # /legislativo/tramit2025/9406.pdf
        parts=path.strip("/").split("/")
        ano=""
        for p in parts:
            if p.startswith("tramit") and p[6:].isdigit(): ano=p[6:]; break
        arq=parts[-1]; numero=os.path.splitext(arq)[0]
        nome=f"Mensagem_{ano}_{numero}.pdf" if ano else arq
        dest=os.path.join(download_dir, nome)
        with open(dest,"wb") as f: f.write(r.content)
        if os.path.getsize(dest)<=0: print("⚠️ PDF 0B:",dest); return None
        print("✅ Baixado:",dest); return dest
    except Exception as e:
        print("⚠️ Exceção ao baixar:",e); return None

# ---- envio via Node
def localizar_sender_js()->Tuple[str,str]:
    for caminho in SENDER_CANDIDATOS:
        if os.path.isfile(caminho): return caminho, os.path.dirname(caminho)
    raise FileNotFoundError("enviar_mensagem.js não encontrado nas pastas padrão.")
def checar_node():
    try:
        out=subprocess.run(["node","-v"],capture_output=True,text=True)
        if out.returncode!=0: raise RuntimeError(out.stderr.strip() or "Node indisponível")
    except FileNotFoundError:
        raise RuntimeError("Instale Node: pkg install -y nodejs-lts")
def enviar_mensagem(numeros:List[str], mensagem:str, caminho_pdf:Optional[str]=None):
    checar_node()
    sender_js, cwd = localizar_sender_js()
    args=["node", sender_js, ",".join(numeros), mensagem]
    if caminho_pdf and os.path.isfile(caminho_pdf) and os.path.getsize(caminho_pdf)>0:
        args.append(caminho_pdf)
    print("▶️ Enviando:", " ".join(shlex.quote(a) for a in args))
    proc=subprocess.Popen(args, cwd=cwd); proc.wait()
    if proc.returncode!=0: print(f"⚠️ enviar_mensagem.js saiu com código {proc.returncode}")

def coletar_mensagens() -> List[Dict]:
    r = requests.get(URL, headers=HEADERS, verify=False, timeout=60)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    # ache o <h3> principal (o grande do expediente)
    
    h3_main = None
    for h3 in soup.find_all("h3"):
        if "expediente" in normalize(h3.get_text(" ", strip=True)):
            h3_main = h3
            break
    if not h3_main:
        # print("⚠️ <h3> principal não encontrado.")
        return []

    itens = []
    # percorre TODOS os <a href> dentro do h3 (pega <b><a> e <a><b>)
    for a in h3_main.find_all("a", href=True):
        titulo = a.get_text(" ", strip=True)  # já pega texto mesmo se houver <b> dentro
        if not titulo:
            continue
        if "mensagem" not in normalize(titulo):
            continue  # só links cujo texto contém 'Mensagem'

        href = a["href"].strip()

        # extrai numero/ano a partir do TEXTO DO LINK
        m = RX_NUM_ANO.search(titulo)
        if not m:
            # às vezes o numero/ano pode estar fora do link; tenta olhar um pai próximo
            parent_text = a.parent.get_text(" ", strip=True) if a.parent else ""
            m = RX_NUM_ANO.search(parent_text)
        if not m:
            # sem numero/ano, ignora
            # print("⚠️ Link com 'Mensagem' mas sem numero/ano:", titulo)
            continue

        numero, ano = m.group(1), m.group(2)

        # descrição: texto logo após o link (ou após o <b> que envolve o link)
        descricao = ""
        # se o <a> está dentro de <b>, descreva após o </b>; senão após o próprio <a>
        base = a.parent if getattr(a.parent, "name", None) == "b" else a

        for sib in base.next_siblings:
            # pare quando chegar em outro bloco forte
            if getattr(sib, "name", None) in {"b", "a"}:
                break
            if getattr(sib, "name", None) == "br":
                continue
            txt = BeautifulSoup(str(sib), "html.parser").get_text(" ", strip=True)
            if txt:
                descricao = txt
                break

        # tenta numero2 (ex.: 9.406 -> 9406) só para registro
        m2 = RX_NUMERO2.search(titulo)
        numero2 = m2.group(1).replace(".", "") if m2 else None

        itens.append({
            "numero": numero,
            "ano": ano,
            "titulo": titulo,
            "descricao": descricao,
            "pdf_url": href,
            "numero2": numero2
        })

    return itens


# ---- loop
def main_loop():
    while True:
        try:
            enviados=enviados_carregar()
            itens=coletar_mensagens()
            print(f"DEBUG: mensagens encontradas = {len(itens)}")
            for it in itens:
                print("-", it["numero"], it["ano"], "|", it["titulo"])
            if not itens: print("Sem itens."); 

            for it in itens:
                num_ano=f"{it['numero']}/{it['ano']}"
                if num_ano in enviados:
                    print("🔁 Já enviado:", num_ano); continue

                mensagem=f"Nova mensagem: {it['titulo']} {it['descricao']}".strip()
                print("MSG:", mensagem)

                caminho_pdf=None
                if it.get("pdf_url"):
                    caminho_pdf=download_pdf(it["pdf_url"], PASTA_PDFS)
                num = NUMEROS_DESTINO
                z = datetime.now()
                zx = z.strftime('%H:%M:%S')
                print(f'Horário: {zx}', end='\r')    
                sleep(4)        
                # Verifica horários específicos e chama a função scan_processos
                if zx >= '16:00:01':
                    print('NUMEROS_DESTINO2')
                    num = NUMEROS_DESTINO2
                if caminho_pdf:
                    enviar_mensagem(num, mensagem, caminho_pdf)
                else:
                    enviar_mensagem(num, mensagem)

                enviados_salvar(num_ano)
                print("✅ Enviado e registrado:", num_ano)
                sleep(2)

        except KeyboardInterrupt:
            print("\nInterrompido."); return
        except Exception as e:
            print("Erro no loop:", e)

        hh=datetime.now().strftime("%H:%M:%S")
        print(f"Horário: {hh} — dormindo 1800s")
        sleep(1800)

if __name__=="__main__":
    main_loop()
