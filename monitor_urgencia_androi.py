


# -*- coding: utf-8 -*-
"""
ALECE PDR — Urgência (Android/Termux)
- Detecta 'urgencia' no conteúdo
- Extrai números (AAAA/AAAA) (tenta Autor, depois Conteúdo; se nada, usa chave sintética)
- Baixa PDF via consulta_plenario.php (POST com leg_id); se não vier PDF, procura link .pdf no HTML
- Envia por WhatsApp via sender_baileys.js
- Dedup em Excel
"""

import os
import re
import shlex
import hashlib
import subprocess
from time import sleep
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import unicodedata

# =========== CONFIG ===========
INTERVALO_SEGUNDOS = 600  # 5 min
URL_BASE_LISTA = "https://www2.al.ce.gov.br/pdr/consultas.php"
URL_PLENARIO   = "https://www2.al.ce.gov.br/pdr/consulta_plenario.php"  # POST com leg_id
PARAM_FIXOS = {
    "opcao": "9",
    "palavra": "",
    "ano_base": "2025",
    "autor": "",
    "numero": "",
    "tipo": "",
    "situacao": "",
    "pg": "publico",
}
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 (KHTML, like Gecko) Chrome Mobile Safari/537.36"
}

# Armazenamento seguro do Termux (aponta p/ sdcard)
BASE_DIR = r"C:\Users\FABIO\OneDrive\Gabinete\site"
ARQ_EXCEL   = os.path.join(BASE_DIR, "requerimentos_urgencia.xlsx")
ABA_EXCEL   = "dados"
PASTA_ANEXO = os.path.join(BASE_DIR, "requerimentos_urgencia")
os.makedirs(PASTA_ANEXO, exist_ok=True)

# Destinatários (TROQUE AQUI)
#NUMEROS_DESTINO = ["558588227227"]
NUMEROS_DESTINO = ["558588227227", "558597159955", "558587262526", "558596195560"]
# Onde procurar o sender (precisa ter node_modules na mesma pasta!)
SENDER_CANDIDATOS = [
    r"C:\Users\FABIO\OneDrive\Gabinete\site\enviar_mensagem.js",                 # ~/bot
    "/storage/emulated/0/Documents/escaner/sender_baileys.js",               # sdcard/Documentos/escaner
]

# =========== Normalização / Filtros ===========
def contem_palavra(texto: str) -> bool:
    """True se 'urgencia' (sem acento) estiver presente (case-insensitive)."""
    if not texto:
        return False
    txt_norm = unicodedata.normalize('NFD', texto)
    txt_sem_acento = ''.join(c for c in txt_norm if unicodedata.category(c) != 'Mn').lower()
    return 'urgencia' in txt_sem_acento

RX_NUMEROS = re.compile(r'\b\d{4}/\d{4}\b')
def extrair_numeros_proposicao(texto: str) -> List[str]:
    return RX_NUMEROS.findall(texto or "")

def verificar_data_menor(texto_data: str) -> bool:
    """True se a data (dd/mm/aaaa) for < data de referência (para encerrar paginação)."""
    m = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", texto_data or "")
    if not m:
        return False
    try:
        data_texto = datetime.strptime(m.group(1), "%d/%m/%Y").date()
    except ValueError:
        return False
    data_referencia = datetime.strptime("05/07/2025", "%d/%m/%Y").date()
    return data_texto < datetime.today().date()

def insertt(num: str) -> str:
    return (num or "").replace("/", "_")

def chave_sintetica(data: str, autor: str, conteudo: str) -> str:
    """Chave única quando não há número 0000/0000."""
    base = f"{data}|{autor}|{conteudo}".encode("utf-8", errors="ignore")
    h = hashlib.sha1(base).hexdigest()[:16]
    return f"K:{h}"

# =========== Excel ===========
def excel_init():
    os.makedirs(BASE_DIR, exist_ok=True)
    if not os.path.exists(ARQ_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = ABA_EXCEL
        ws.append(["Id"])  # número 0000/0000 ou K:<hash>
        wb.save(ARQ_EXCEL)

def carregar_existentes() -> set:
    excel_init()
    wb = load_workbook(ARQ_EXCEL)
    ws = wb[ABA_EXCEL]
    return {str(row[0].value).strip() for row in ws.iter_rows(min_row=2) if row[0].value}

def salvar_novo(ident: str):
    wb = load_workbook(ARQ_EXCEL)
    ws = wb[ABA_EXCEL]
    ws.append([ident])
    wb.save(ARQ_EXCEL)

# =========== Sender (Node) ===========
def localizar_sender() -> Tuple[str, str]:
    for caminho in SENDER_CANDIDATOS:
        if os.path.isfile(caminho):
            return caminho, os.path.dirname(caminho)
    raise FileNotFoundError("sender_baileys.js não encontrado nas pastas padrão.")

def checar_node():
    try:
        out = subprocess.run(["node", "-v"], capture_output=True, text=True)
        if out.returncode != 0:
            raise RuntimeError(out.stderr.strip() or "Node indisponível")
    except FileNotFoundError:
        raise RuntimeError("Node.js não encontrado. Instale com: pkg install -y nodejs-lts")

def checar_dependencias_sender(cwd: str):
    nm = os.path.join(cwd, "node_modules")
    pj = os.path.join(cwd, "package.json")
    if not os.path.isdir(nm) or not os.path.isfile(pj):
        raise RuntimeError(
            f"Dependências do sender ausentes em {cwd}. "
            "Rode: npm init -y && npm i @whiskeysockets/baileys qrcode-terminal "
            "(em /storage/... use --no-bin-links)."
        )

def chamar_sender(numeros: List[str], msg: str, caminho_pdf: Optional[str] = None):
    checar_node()
    sender_js, sender_cwd = localizar_sender()
    checar_dependencias_sender(sender_cwd)
    args = ["node", sender_js, ",".join(numeros), msg]
    if caminho_pdf:
        args.append(caminho_pdf)
    print("▶️ Executando sender:", " ".join(shlex.quote(a) for a in args))
    print("📂 CWD:", sender_cwd)
    proc = subprocess.Popen(args, cwd=sender_cwd)
    proc.wait()
    if proc.returncode != 0:
        print(f"⚠️ sender saiu com código {proc.returncode}")

# =========== Download via PLENÁRIO (POST com leg_id) ===========
def salvar_stream_em_pdf(resp: requests.Response, destino: str) -> Optional[str]:
    tmp = destino + ".part"
    with open(tmp, "wb") as f:
        for chunk in resp.iter_content(8192):
            if chunk:
                f.write(chunk)
    os.replace(tmp, destino)
    if os.path.getsize(destino) <= 0:
        try: os.remove(destino)
        except: pass
        return None
    return destino

def baixar_via_plenario(leg_id: str, nome_base: str) -> Optional[str]:
    """
    Baixa o PDF via POST em consulta_plenario.php.
    - Se o servidor devolver PDF direto, salva.
    - Se devolver HTML, procura <a href="*.pdf"> e baixa o primeiro.
    """
    if not leg_id:
        return None
    try:
        data = {"leg_id": str(leg_id), "pg": "publico", "visualizar": "Visualizar"}
        print(f"⇣ POST {URL_PLENARIO} leg_id={leg_id}")
        with requests.post(URL_PLENARIO, data=data, headers=HEADERS, timeout=60,
                           allow_redirects=True, stream=True) as r:
            r.raise_for_status()
            ctype = (r.headers.get("Content-Type") or "").lower()

            # Caso 1: já veio PDF
            if "pdf" in ctype:
                destino = os.path.join(PASTA_ANEXO, f"{nome_base}.pdf")
                ok = salvar_stream_em_pdf(r, destino)
                if ok:
                    print(f"  ✓ PDF (direto) salvo: {ok}")
                    return ok
                print("  ✖️ Falha ao salvar PDF direto")
                return None

            # Caso 2: veio HTML — procurar link .pdf
            html = r.text
            soup = BeautifulSoup(html, "html.parser")
            link = None
            for a in soup.select('a[href]'):
                href = a.get("href", "")
                if ".pdf" in href.lower():
                    link = urljoin(URL_PLENARIO, href)
                    break
            if not link:
                print("  ⚠️ Nenhum link .pdf encontrado no HTML do plenário")
                return None

            print("  → Link PDF encontrado:", link)
            with requests.get(link, headers=HEADERS, timeout=60, stream=True) as rb:
                rb.raise_for_status()
                if "pdf" not in (rb.headers.get("Content-Type") or "").lower():
                    print("  ⚠️ Link não retornou PDF")
                    return None
                destino = os.path.join(PASTA_ANEXO, f"{nome_base}.pdf")
                ok = salvar_stream_em_pdf(rb, destino)
                if ok:
                    print(f"  ✓ PDF (link) salvo: {ok}")
                    return ok
                print("  ✖️ Falha ao salvar PDF do link")
                return None
    except Exception as e:
        print("  ✖️ baixar_via_plenario falhou:", e)
        return None

# =========== Raspagem ===========
def baixar_pagina(pagina: int) -> str:
    params = dict(PARAM_FIXOS)
    params["pagina"] = str(pagina)
    r = requests.get(URL_BASE_LISTA, params=params, headers=HEADERS, timeout=60)
    r.raise_for_status()
    return r.text

def extrair_leg_ids(soup: BeautifulSoup) -> List[str]:
    return [inp.get("value","") for inp in soup.select('input[name="leg_id"]')]

def parse_linhas(html: str) -> List[Dict]:
    soup = BeautifulSoup(html, "html.parser")
    leg_ids = extrair_leg_ids(soup)
    resultados = []
    trs = soup.select("table tr")
    idx_leg = 0
    for tr in trs:
        spans = tr.select("td span")
        if len(spans) < 3:
            continue
        data_txt  = spans[0].get_text(strip=True)
        autor_txt = spans[1].get_text(strip=True)
        cont_txt  = spans[2].get_text(strip=True)
        leg_id = leg_ids[idx_leg] if idx_leg < len(leg_ids) else None
        idx_leg += 1
        resultados.append({"data": data_txt, "autor": autor_txt, "conteudo": cont_txt, "leg_id": leg_id})
    return resultados

# =========== Loop principal ===========
def main_loop():
    while True:
        try:
            existentes = carregar_existentes()
            print(">>> Iniciando varredura de urgência...")
            pagina = 1
            encerrar = False

            while True:
                try:
                    html = baixar_pagina(pagina)
                except Exception as e:
                    print(f"Falha ao baixar página {pagina}:", e)
                    break

                linhas = parse_linhas(html)
                if not linhas:
                    print("Sem linhas nesta página.")
                    break

                for linha in linhas:
                    data = linha["data"]
                    autor = linha["autor"]
                    conteudo = linha["conteudo"]
                    leg_id = linha.get("leg_id")

                    # parar paginação quando a linha estiver antes da data de referência
                    if verificar_data_menor(data):
                        encerrar = True
                        break

                    # filtra por 'urgencia' no conteúdo
                    if not contem_palavra(conteudo):
                        continue

                    # tenta número 0000/0000 no Autor, depois no Conteúdo
                    nums = extrair_numeros_proposicao(autor) or extrair_numeros_proposicao(conteudo)

                    # se não há número, crie chave sintética (para não perder o envio)
                    ids_para_enviar = nums if nums else [chave_sintetica(data, autor, conteudo)]
                    if not nums:
                        print("ℹ️ Sem número 0000/0000 — usando chave sintética:", ids_para_enviar[0])

                    # dedupe
                    nova = False
                    for ident in ids_para_enviar:
                        if ident not in existentes:
                            salvar_novo(ident)
                            existentes.add(ident)
                            nova = True
                            print(f"📌 Nova ocorrência: {ident}")
                        else:
                            print(f"🔁 Já registrada: {ident}")
                    if not nova:
                        continue

                    # mensagem
                    mensagem = f"{data}\n\n{autor}\n\n{conteudo}".strip()
                    print("MSG:", mensagem)

                    # nome-base do arquivo (para salvar com sentido)
                    nome_base = insertt(ids_para_enviar[0]) if "/" in ids_para_enviar[0] else ids_para_enviar[0].replace("K:", "K_")

                    # Baixar PDF via plenário (POST com leg_id)
                    caminho_pdf = baixar_via_plenario(leg_id, nome_base) if leg_id else None

                    # Envio
                    if caminho_pdf and os.path.isfile(caminho_pdf) and os.path.getsize(caminho_pdf) > 0:
                        print("→ enviando com PDF:", caminho_pdf)
                        chamar_sender(NUMEROS_DESTINO, mensagem, caminho_pdf)
                    else:
                        if caminho_pdf:
                            print("⚠️ PDF inválido/0B — envio só texto:", caminho_pdf)
                        chamar_sender(NUMEROS_DESTINO, mensagem)

                if encerrar:
                    print("↩️ Encontrou data anterior à referência — encerrando paginação.")
                    break

                pagina += 1

        except KeyboardInterrupt:
            print("\nInterrompido pelo usuário.")
            return
        except Exception as e:
            print("Erro no loop:", e)

        hh = datetime.now().strftime("%H:%M:%S")
        print(f"Horário: {hh} — dormindo {INTERVALO_SEGUNDOS}s")
        sleep(INTERVALO_SEGUNDOS)

if __name__ == "__main__":
    main_loop()
