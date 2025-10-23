# -*- coding: utf-8 -*-
"""
Raspa a listagem pública (ALECE PDR) com Requests+BeautifulSoup,
detecta 'urgencia' no conteúdo, extrai numeros (AAAA/AAAA) do autor,
( opcional ) baixa anexo via anexo.baixar(leg_id, numero_formatado),
envia por WhatsApp via sender_baileys.js,
e registra em Excel para não duplicar.

Termux (uma vez):
  pkg update -y
  pkg install -y python nodejs-lts
  pip install --upgrade pip
  pip install requests beautifulsoup4 openpyxl

Ativar acesso ao armazenamento:
  termux-setup-storage
"""

import os
import re
import sys
import json
import shlex
import subprocess
from time import sleep
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import unicodedata

# =========== CONFIG ===========
INTERVALO_SEGUNDOS = 300  # 5 min
URL_BASE = "https://www2.al.ce.gov.br/pdr/consultas.php"
PARAM_FIXOS = {
    "opcao": "9",
    "palavra": "",
    "ano_base": "2025",
    "autor": "",
    "numero": "",
    "tipo": "",
    "situacao": "",
    "pg": "publico",
    # pagina=1,2,3,... será adicionada no loop
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 (KHTML, like Gecko) Chrome Mobile Safari/537.36"
}

# Use um caminho garantido do Termux para o armazenamento compartilhado (=/sdcard)
BASE_DIR = "/data/data/com.termux/files/home/storage/documents/escaner"
os.makedirs(BASE_DIR, exist_ok=True)
ARQ_EXCEL  = os.path.join(BASE_DIR, "requerimentos_urgencia.xlsx")
ABA_EXCEL  = "dados"
PASTA_ANEXO = os.path.join(BASE_DIR, "mensagens")
os.makedirs(PASTA_ANEXO, exist_ok=True)

# Para quem enviar
#NUMEROS_DESTINO = ["558588227227"]  # <- TROQUE PELO(S) SEU(S) NÚMERO(S)
NUMEROS_DESTINO = ["558588227227", "558597159955", "558587262526", "558596195560"]
# Onde procurar o sender (precisa ter node_modules na mesma pasta!)
SENDER_CANDIDATOS = [
    "/data/data/com.termux/files/home/bot/sender_baileys.js",                 # ~/bot
    "/storage/emulated/0/Documents/escaner/sender_baileys.js",               # sdcard/Documentos/escaner
]

# =========== Funções de normalização/filtro ===========
def contem_palavra(texto: str) -> bool:
    if not texto:
        return False
    txt_norm = unicodedata.normalize('NFD', texto)
    txt_sem_acento = ''.join(c for c in txt_norm if unicodedata.category(c) != 'Mn').lower()
    return 'urgencia' in txt_sem_acento

RX_NUMEROS = re.compile(r'\b\d{4}/\d{4}\b')
def extrair_numeros_proposicao(texto: str):
    return RX_NUMEROS.findall(texto or "")

def verificar_data_menor(texto_data: str) -> bool:
    """
    Retorna True se a data no texto (dd/mm/aaaa) for < hoje (para encerrar paginação).
    """
    m = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", texto_data or "")
    if not m:
        return False
    try:
        data_texto = datetime.strptime(m.group(1), "%d/%m/%Y").date()
    except ValueError:
        return False
    data_referencia = datetime.strptime("05/07/2025", "%d/%m/%Y").date()
    return data_texto < data_referencia

def insertt(num: str) -> str:
    return (num or "").replace("/", "_")

# =========== Excel helpers ===========
def excel_init():
    if not os.path.exists(ARQ_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = ABA_EXCEL
        ws.append(["Numero"])
        wb.save(ARQ_EXCEL)

def carregar_existentes() -> set:
    excel_init()
    wb = load_workbook(ARQ_EXCEL)
    ws = wb[ABA_EXCEL]
    return {str(row[0].value).strip() for row in ws.iter_rows(min_row=2) if row[0].value}

def salvar_novo(numero: str):
    wb = load_workbook(ARQ_EXCEL)
    ws = wb[ABA_EXCEL]
    ws.append([numero])
    wb.save(ARQ_EXCEL)

# =========== Sender (Node) ===========
def localizar_sender() -> Tuple[str, str]:
    for caminho in SENDER_CANDIDATOS:
        if os.path.isfile(caminho):
            return caminho, os.path.dirname(caminho)
    raise FileNotFoundError(
        "Não encontrei 'sender_baileys.js' nos caminhos:\n- " +
        "\n- ".join(SENDER_CANDIDATOS)
    )

def checar_node():
    try:
        out = subprocess.run(["node", "-v"], capture_output=True, text=True)
        if out.returncode != 0:
            raise RuntimeError(out.stderr.strip() or "Node indisponível")
    except FileNotFoundError:
        raise RuntimeError("Node.js não encontrado. Instale com: pkg install -y nodejs-lts")

def checar_dependencias_sender(cwd: str):
    nm = os.path.join(cwd, "node_modules")
    pj = os.path.join(cwd, "package.json")
    if not os.path.isdir(nm) or not os.path.isfile(pj):
        raise RuntimeError(
            f"Dependências do sender ausentes em:\n{cwd}\n"
            "Entre nessa pasta e rode:\n"
            "  npm init -y\n"
            "  npm i @whiskeysockets/baileys qrcode-terminal\n"
            "(se for no /storage/emulated/0/, use --no-bin-links)"
        )

def chamar_sender(numeros: List[str], msg: str, caminho_pdf: Optional[str] = None):
    checar_node()
    sender_js, sender_cwd = localizar_sender()
    checar_dependencias_sender(sender_cwd)

    args = ["node", sender_js, ",".join(numeros), msg]
    if caminho_pdf:
        args.append(caminho_pdf)

    print("▶️ Executando sender:", " ".join(shlex.quote(a) for a in args))
    print("📂 CWD:", sender_cwd)
    # Executa mostrando logs do Node (QR, Estado: open, etc.)
    proc = subprocess.Popen(args, cwd=sender_cwd)
    proc.wait()
    if proc.returncode != 0:
        print(f"⚠️ sender finalizou com código {proc.returncode}")

# =========== (Opcional) baixar anexo via anexo.baixar ===========
def tentar_baixar_anexo(leg_id: Optional[str], numero_fmt: str) -> Optional[str]:
    """
    Se o módulo anexo.baixar existir, usa-o. Caso contrário, pula o PDF.
    """
    if not leg_id:
        return None
    try:
        from anexo import baixar
    except Exception:
        return None  # sem módulo anexo
    try:
        caminho = baixar(leg_id, numero_fmt)
        if caminho and os.path.isfile(caminho):
            # se veio em pasta diferente, copie para PASTA_ANEXO (opcional)
            return caminho
    except Exception as e:
        print("Falha no anexo.baixar:", e)
    return None

# =========== Raspagem (Requests + BS4) ===========
def baixar_pagina(pagina: int) -> str:
    params = dict(PARAM_FIXOS)
    params["pagina"] = str(pagina)
    r = requests.get(URL_BASE, params=params, headers=HEADERS, timeout=60)
    r.raise_for_status()
    return r.text

def extrair_leg_ids(soup: BeautifulSoup) -> List[str]:
    # Ordem dos inputs deve bater com a ordem visual
    return [inp.get("value","") for inp in soup.select('input[name="leg_id"]')]

def parse_linhas(html: str) -> List[Dict]:
    soup = BeautifulSoup(html, "html.parser")  # evita necessidade do lxml
    leg_ids = extrair_leg_ids(soup)
    resultados = []

    trs = soup.select("table tr")
    # A página parece alternar linhas de conteúdo; vamos varrer todas
    idx_leg = 0
    for tr in trs:
        spans = tr.select("td span")
        if len(spans) < 3:
            continue

        data_txt  = spans[0].get_text(strip=True)  # <span><strong>DATA</strong></span>
        autor_txt = spans[1].get_text(strip=True)  # <span><strong>AUTOR</strong></span>
        cont_txt  = spans[2].get_text(strip=True)  # <span>CONTEUDO</span> (geralmente)

        # pega um leg_id se houver na sequência
        leg_id = leg_ids[idx_leg] if idx_leg < len(leg_ids) else None
        idx_leg += 1

        resultados.append({
            "data": data_txt,
            "autor": autor_txt,
            "conteudo": cont_txt,
            "leg_id": leg_id
        })
    return resultados

# =========== Loop principal ===========
def main_loop():
    while True:
        try:
            existentes = carregar_existentes()
            print(">>> Iniciando varredura de urgência...")
            pagina = 1
            encerrar = False

            while True:
                try:
                    html = baixar_pagina(pagina)
                except Exception as e:
                    print(f"Falha ao baixar página {pagina}:", e)
                    break

                linhas = parse_linhas(html)
                if not linhas:
                    print("Sem linhas nesta página.")
                    break

                for linha in linhas:
                    data = linha["data"]
                    autor = linha["autor"]
                    conteudo = linha["conteudo"]
                    leg_id = linha.get("leg_id")

                    # parar paginação quando o registro for mais antigo que hoje
                    if verificar_data_menor(data):
                        encerrar = True
                        break

                    # filtra por 'urgencia' no conteúdo
                    if not contem_palavra(conteudo):
                        continue

                    nums = extrair_numeros_proposicao(autor)
                    if not nums:
                        continue

                    nova_detectada = False
                    for numero in nums:
                        if numero not in existentes:
                            print(f"📌 Nova proposição detectada: {numero}")
                            salvar_novo(numero)
                            existentes.add(numero)
                            nova_detectada = True
                        else:
                            print(f"🔁 Já registrada: {numero}")

                    if not nova_detectada:
                        continue

                    # monta mensagem
                    mensagem = f"{data}\n\n{autor}\n\n{conteudo}".strip()

                    # tenta baixar anexo via anexo.baixar (se existir)
                    caminho_pdf = None
                    try:
                        numero_fmt = insertt(nums[0])
                        caminho_pdf = tentar_baixar_anexo(leg_id, numero_fmt)
                    except Exception as e:
                        print("Erro ao tentar baixar anexo:", e)
                        caminho_pdf = None

                    # envia
                    if caminho_pdf:
                        chamar_sender(NUMEROS_DESTINO, mensagem, caminho_pdf)
                    else:
                        chamar_sender(NUMEROS_DESTINO, mensagem)

                if encerrar:
                    print("↩️ Encontrou data anterior a hoje — encerrando paginação.")
                    break

                pagina += 1

        except KeyboardInterrupt:
            print("\nInterrompido pelo usuário.")
            return
        except Exception as e:
            print("Erro no loop:", e)

        hh = datetime.now().strftime("%H:%M:%S")
        print(f"Horário: {hh} — dormindo {INTERVALO_SEGUNDOS}s")
        sleep(INTERVALO_SEGUNDOS)

# if __name__ == "__main__":
#     main_loop()

import os
import re
import sys
import json
import shlex
import subprocess
from time import sleep
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import unicodedata

# =========== CONFIG ===========
URL = "https://www.al.ce.gov.br/legislativo/expediente"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 (KHTML, like Gecko) Chrome Mobile Safari/537.36"
}

RX_PUBLICACAO  = re.compile(r"\b(\d{2,5})/(\d{4})\b\s*-\s*.+")   # ex: 823/2025 - ...
RX_PUBLICACAO2 = re.compile(r"\b(\d{1,2}\.\d{3})\b\s*-\s*.+")   # ex: 9.402 - ...

BASE_DIR = r"C:\Users\FABIO\OneDrive\Gabinete\site"
os.makedirs(BASE_DIR, exist_ok=True)

ARQ_EXCEL  = os.path.join(BASE_DIR, "mensagens_encontradas.xlsx")
ABA_EXCEL  = "encontradas"
PASTA_PDFS = os.path.join(BASE_DIR, "mensagens")
os.makedirs(PASTA_PDFS, exist_ok=True)

NUMEROS_DESTINO = ["558588227227"]
# NUMEROS_DESTINO = ["558588227227", "558597159955", "558587262526"]
SENDER_CANDIDATOS = [
    r"C:\Users\FABIO\OneDrive\Gabinete\site\enviar_mensagem.js",
    "/storage/emulated/0/Documents/escaner/sender_baileys.js",
]

# =========== Funções auxiliares ===========
def contem_palavra(txt: str) -> bool:
    if not txt:
        return False
    t = unicodedata.normalize("NFD", txt)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn").lower()
    return "leitura" in t

def contem_palavra2(txt: str) -> bool:
    if not txt:
        return False
    t = unicodedata.normalize("NFD", txt)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn").lower()
    return "mensagem" in t

# =========== Excel helpers ===========
def excel_init():
    p = Path(ARQ_EXCEL)
    if p.exists():
        wb = load_workbook(ARQ_EXCEL)
        ws = wb[ABA_EXCEL] if ABA_EXCEL in wb.sheetnames else wb.create_sheet(ABA_EXCEL)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = ABA_EXCEL
        ws.append(["numero"])
        wb.save(ARQ_EXCEL)
    return wb, ws

def carregar_numeros_enviados() -> set:
    wb, ws = excel_init()
    enviados = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if not row:
            continue
        v = (row[0] or "").strip()
        if v:
            enviados.add(v)
    return enviados

def salvar_numero_enviado(num_ano: str):
    wb, ws = excel_init()
    ws.append([num_ano])
    wb.save(ARQ_EXCEL)

# =========== Localizar sender ===========
def localizar_sender() -> Tuple[str, str]:
    for caminho in SENDER_CANDIDATOS:
        if os.path.isfile(caminho):
            return caminho, os.path.dirname(caminho)
    raise FileNotFoundError(
        "Não encontrei 'sender_baileys.js' nos caminhos:\n- " +
        "\n- ".join(SENDER_CANDIDATOS)
    )

def checar_node():
    try:
        out = subprocess.run(["node", "-v"], capture_output=True, text=True)
        if out.returncode != 0:
            raise RuntimeError(out.stderr.strip() or "node indisponível")
    except FileNotFoundError:
        raise RuntimeError("Node.js não encontrado. Instale com: pkg install -y nodejs-lts")

def checar_dependencias_sender(cwd: str):
    nm = os.path.join(cwd, "node_modules")
    pj = os.path.join(cwd, "package.json")
    if not os.path.isdir(nm) or not os.path.isfile(pj):
        raise RuntimeError(
            f"Dependências do sender ausentes em:\n{cwd}\n"
            "npm init -y && npm i @whiskeysockets/baileys qrcode-terminal"
        )

# =========== Download de PDF ===========
def _pdf_candidates(ano: str, numero2: str) -> list:
    ano = str(ano)
    yy = ano[-2:]
    n = str(numero2 or "").strip().replace(".", "")
    nums = {n}
    for w in (3, 4, 5):
        nums.add(n.zfill(w))
    prefixes = ["", "pl", "plc", "pdl", "msg", "au"]
    urls = []
    for num in nums:
        urls.append(f"https://www2.al.ce.gov.br/legislativo/tramit{ano}/{num}.pdf")
        urls.append(f"https://www2.al.ce.gov.br/legislativo/tramit{ano}/{num}_{yy}.pdf")
        for p in prefixes[1:]:
            urls.append(f"https://www2.al.ce.gov.br/legislativo/tramit{ano}/{p}{num}_{yy}.pdf")
    seen = set(); uniq = []
    for u in urls:
        if u not in seen:
            uniq.append(u); seen.add(u)
    return uniq

def try_download_first_pdf(urls: list, pasta_dest: str, headers: dict) -> Optional[str]:
    for url in urls:
        try:
            print("⇣ Tentando:", url)
            with requests.get(url, headers=headers, timeout=60, stream=True) as r:
                r.raise_for_status()
                ctype = (r.headers.get("Content-Type") or "").lower()
                if "pdf" not in ctype:
                    continue
                nome = url.rstrip("/").split("/")[-1]
                dest = os.path.join(pasta_dest, nome)
                tmp = dest + ".part"
                with open(tmp, "wb") as f:
                    for chunk in r.iter_content(8192):
                        if chunk:
                            f.write(chunk)
                os.replace(tmp, dest)
                if os.path.getsize(dest) > 0:
                    print(f"✓ PDF salvo: {dest}")
                    return dest
        except Exception:
            continue
    return None

# =========== Raspar página ===========
def get_text_nodes_outside_b(h3) -> str:
    parts = []
    for node in h3.children:
        if getattr(node, "name", None) is None:
            s = str(node).strip()
            if s:
                parts.append(s)
    return " ".join(parts).strip()

def raspar_itens() -> List[Dict]:
    r = requests.get(URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    itens = []
    for h3 in soup.select("main div div div h3"):
        titulo_b = (h3.b.get_text(" ", strip=True) if h3.b else "").strip()
        texto_solto = get_text_nodes_outside_b(h3)
        if not titulo_b:
            continue
        if contem_palavra(titulo_b):
            itens.append({"tipo": "leitura","titulo_b": titulo_b,"texto_solto": texto_solto})
            continue
        m = RX_PUBLICACAO.search(titulo_b)
        if not m:
            continue
        numero, ano = m.group(1), m.group(2)
        if not contem_palavra2(titulo_b):
            continue
        m2 = RX_PUBLICACAO2.search(titulo_b)
        numero2 = m2.group(1).replace(".", "") if m2 else None
        itens.append({
            "tipo": "mensagem","titulo_b": titulo_b,"texto_solto": texto_solto,
            "numero": numero,"ano": ano,"numero2": numero2
        })
    return itens

# =========== Envio ===========
def chamar_sender(numeros: List[str], msg: str, caminho_pdf: Optional[str] = None):
    checar_node()
    sender_js, sender_cwd = localizar_sender()
    checar_dependencias_sender(sender_cwd)
    args = ["node", sender_js, ",".join(numeros), msg]
    if caminho_pdf: args.append(caminho_pdf)
    print("▶️ Enviando:", msg)
    subprocess.run(args, cwd=sender_cwd)

# =========== Main ===========
def main_loop():
    while True:
        try:
            enviados = carregar_numeros_enviados()
            itens = raspar_itens()
            for it in itens:
                if it["tipo"] != "mensagem":
                    continue
                num_ano = f"{it['numero']}/{it['ano']}"
                if num_ano in enviados:
                    continue
                mensagem = "Nova mensagem: " + it["titulo_b"] + " " + it["texto_solto"]
                print("MSG:", mensagem)
                # checagem de data
                if data_menor_que_referencia(it["titulo_b"] + " " + it["texto_solto"]):
                    print("⏭️ Ignorando item anterior à data de referência.")
                    continue
                caminho_pdf = None
                if it.get("ano") and it.get("numero2"):
                    cands = _pdf_candidates(it["ano"], it["numero2"])
                    caminho_pdf = try_download_first_pdf(cands, PASTA_PDFS, HEADERS)
                if caminho_pdf:
                    chamar_sender(NUMEROS_DESTINO, mensagem, caminho_pdf)
                else:
                    chamar_sender(NUMEROS_DESTINO, mensagem)
                salvar_numero_enviado(num_ano)
                print("✅ Enviado e registrado:", num_ano)
        except KeyboardInterrupt:
            return
        except Exception as e:
            print("Erro:", e)
        print("Dormindo 3600s...")
        sleep(3600)

if __name__ == "__main__":
    main_loop()




